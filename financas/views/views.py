from financas.serializer.dto.serializers import FinancasSerializer
from financas.models.entity.financas_model import Financas
from rest_framework.response import Response
from rest_framework import generics
from rest_framework import status
from dotenv import load_dotenv
from tabulate import tabulate
import openpyxl
import requests
import logging
import json
import os

load_dotenv()

# Defina o nível de logging como INFO ou DEBUG
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

notion_token = os.getenv("NOTION_TOKEN")
headers = {
    'Authorization': f"Bearer {notion_token}",
    'Content-Type': 'application/json',
    'Notion-Version': '2022-02-22'
}
banco_notion = os.getenv("ID_DO_BANCO")

if not notion_token or not banco_notion:
    raise ValueError(
        "Token de acesso ao Notion ou ID do banco não encontrado no arquivo .env"
    )

# Buscar dados no Notion


def get_data_from_notion():
    search_params = {"filter": {"value": "page", "property": "object"}}
    search_response = requests.post(
        f'https://api.notion.com/v1/search',
        json=search_params, headers=headers)
    data = json.dumps(search_response.json(), indent=4, ensure_ascii=False)
    print(data)


# get_data_from_notion()

# obter dados das propriedades do Notion

def get_database_properties():
    url = f"https://api.notion.com/v1/databases/{banco_notion}"
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        properties = response.json().get("properties", {})
        print(json.dumps(properties, indent=4, ensure_ascii=False))
    else:
        print(f"Erro ao buscar propriedades do banco de dados: {
              response.text}")

# get_database_properties()


def save_or_update_in_sheet(notion_data):
    logger.debug(f"Chamando save_or_update_in_sheet com dados: {notion_data}")
    file_path = "./planilhas/Finanças.xlsx"
    sheet_name = "Finanças"

    try:
        # Cria um novo workbook se não existir
        if not os.path.exists(file_path):
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = sheet_name
            worksheet.append(["Nome", "Entradas", "Saídas",
                             "Saldo", "Notion Page ID"])
            workbook.save(file_path)

        # Carrega o workbook existente
        logger.debug("Tentando carregar a planilha")
        workbook = openpyxl.load_workbook(file_path)
        logger.debug("Planilha carregada com sucesso")

        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.create_sheet(title=sheet_name)
            worksheet.append(["Nome", "Entradas", "Saídas",
                             "Saldo", "Notion Page ID"])

        # Atualiza ou adiciona a linha na planilha
        id_exists = False
        logger.debug(
            "Iniciando verificação de existência do Notion Page ID na planilha")

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=False):
            notion_id = row[4].value
            if notion_id == notion_data["notion_page_id"]:
                logger.debug(f"ID {notion_id} encontrado na linha {
                             row[0].row}, atualizando dados")
                row[0].value = notion_data["nome"]
                row[1].value = notion_data["entradas"]
                row[2].value = notion_data["saidas"]
                row[3].value = notion_data["saldo"]
                id_exists = True
                updated_row = [notion_data["nome"], notion_data["entradas"],
                               notion_data["saidas"], notion_data["saldo"], notion_data["notion_page_id"]]
                break

        if not id_exists:
            logger.debug("ID não encontrado, adicionando nova linha")
            worksheet.append([
                notion_data["nome"],
                notion_data["entradas"],
                notion_data["saidas"],
                notion_data["saldo"],
                notion_data["notion_page_id"]
            ])
            updated_row = [notion_data["nome"], notion_data["entradas"],
                           notion_data["saidas"], notion_data["saldo"], notion_data["notion_page_id"]]

        workbook.save(file_path)
        logger.info(f"Planilha atualizada e salva em: {file_path}")

        # Log do conteúdo atualizado da linha
        logger.info("Conteúdo da linha atualizada:")
        table_headers = ["Nome", "Entradas",
                         "Saídas", "Saldo", "Notion Page ID"]
        logger.info(
            tabulate([updated_row], headers=table_headers, tablefmt="grid"))

    except Exception as e:
        logger.error(f"Erro ao atualizar ou salvar a planilha: {e}")
        raise


# Excluir dados do Excel

def delete_from_sheet(notion_page_id):
    file_path = "./planilhas/Finanças.xlsx"
    sheet_name = "Finanças"

    if not os.path.exists(file_path):
        logger.warning("O arquivo de planilha não foi encontrado.")
        return

    workbook = openpyxl.load_workbook(file_path)

    if sheet_name not in workbook.sheetnames:
        logger.warning("A planilha '%s' não existe no arquivo.", sheet_name)
        return

    worksheet = workbook[sheet_name]

    # Encontre e exclua a linha com o Notion Page ID correspondente
    rows_to_delete = []
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=False):
        if row[4].value == notion_page_id:
            rows_to_delete.append(row[0].row)

    # Excluir as linhas encontradas na ordem inversa
    for row_idx in reversed(rows_to_delete):
        worksheet.delete_rows(row_idx)

    workbook.save(file_path)
    logger.info(f"Linhas com Notion Page ID {
                notion_page_id} excluídas da planilha.")


class FinancasCreateView(generics.CreateAPIView):
    serializer_class = FinancasSerializer

    def get_fields(self):
        fields = [field.name for field in Financas._meta.get_fields()]
        fields.remove('notion_page_id')
        return fields

    def create(self, request, *args, **kwargs):
        try:
            data = request.data
            url = "https://api.notion.com/v1/pages"
            payload = {
                "parent": {"database_id": banco_notion},
                "properties": {
                    "umtC": {  # ID para "Entradas"
                        "number": data.get("entradas", 0)
                    },
                    "wFRQ": {  # ID para "Saídas "
                        "number": data.get("saidas", 0)
                    },
                    "~Pfs": {  # ID para "Saldo"
                        "number": data.get("saldo", 0)
                    },
                    "title": {  # ID para "Nome"
                        "title": [
                            {"text": {"content": data.get("nome", "")}}
                        ]
                    }
                }
            }

            # Esses ID'de entradas facilitam o mapeamento  das informações vindas do Notion,
            # considerando que cada campo da tabela no Notion tem um ID

            response = requests.post(url, json=payload, headers=headers)

            if response.status_code in [200, 201]:
                # Salvar a tarefa no banco de dados
                notion_id = response.json()["id"]
                notion = Financas.objects.create(
                    nome=data["nome"],
                    entradas=data["entradas"],
                    saidas=data["saidas"],
                    saldo=data["saldo"],
                    notion_page_id=notion_id
                )

                # Serializar os dados do objeto criado para retorno das respostas
                serializer = self.get_serializer(notion)
                serialized_notion = serializer.data

                # Atualizar a planilha
                notion_data = {
                    "notion_page_id": notion_id,
                    "nome": data["nome"],
                    "entradas": data["entradas"],
                    "saidas": data["saidas"],
                    "saldo": data["saldo"],
                }
                save_or_update_in_sheet(notion_data)

                # Registrar os dados no log em formato JSON
                logger.info(json.dumps(serialized_notion,
                            indent=4, ensure_ascii=False))

                # Retornar os dados criados na resposta
                return Response({"message": "Finança criada com sucesso", "data": serialized_notion}, status=status.HTTP_201_CREATED)

            else:
                # Se a resposta do Notion não for 200 ou 201, tratar o erro...
                logger.error(
                    "Erro ao criar página no Notion: %s", response.text)
                return Response({"message": "Erro ao criar finança no Notion"}, status=status.HTTP_400_BAD_REQUEST)

        except Exception as erro:
            logger.error("Erro ao criar Notion: %s", erro)
            return Response({"message": "Erro ao criar finança"}, status=status.HTTP_400_BAD_REQUEST)


class FinancasUpdateView(generics.UpdateAPIView):
    queryset = Financas.objects.all()
    serializer_class = FinancasSerializer
    lookup_field = 'pk'

    def update(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            data = request.data

            # Atualizar no banco de dados
            serializer = self.get_serializer(instance, data=data, partial=True)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            updated_notion = serializer.instance

            entradas = float(data.get("entradas", 0))
            saidas = float(data.get("saidas", 0))
            saldo = float(data.get("saldo", 0))

            # Atualizar a página do Notion

            notion_update_data = {
                "properties": {
                    "umtC": {
                        "number": entradas
                    },
                    "wFRQ": {
                        "number": saidas
                    },
                    "~Pfs": {
                        "number": saldo
                    },
                    "title": {
                        "title": [
                            {"text": {"content": data.get("nome", "")}}
                        ]
                    }
                }
            }
            url = f"https://api.notion.com/v1/pages/{
                updated_notion.notion_page_id}"
            response = requests.patch(
                url, json=notion_update_data, headers=headers)

            if response.status_code not in [200, 202]:
                logger.error(
                    "Erro ao atualizar a página no Notion: %s", response.text)
                raise Exception("Erro ao atualizar a página no Notion")

            notion_data = {
                "notion_page_id": updated_notion.notion_page_id,
                "nome":  data.get("nome", updated_notion.nome),
                "entradas": data.get("entradas",  updated_notion.entradas),
                "saidas": data.get("saidas",  updated_notion.saidas),
                "saldo": data.get("saldo",  updated_notion.saldo),
            }
            save_or_update_in_sheet(notion_data)

            # Serializar o objeto atualizado e os dados da resposta
            serialized_notion = FinancasSerializer(updated_notion).data
            response_data = {
                "message": "Finança atualizada com sucesso", "data": serializer.data}

            # Registrar os dados no log em formato JSON
            logger.info("Dados Atualizados: %s", json.dumps(
                serialized_notion, indent=4, ensure_ascii=False))
            logger.info("Resposta: %s", json.dumps(
                response_data, indent=4, ensure_ascii=False))

            return Response(response_data, status=status.HTTP_200_OK)
        except Exception as erro:
            logger.error("Erro ao atualizar Notion: %s", erro)
            return Response({"message": "Erro ao atualizar tarefa"}, status=status.HTTP_400_BAD_REQUEST)


class FinancasListView(generics.ListAPIView):
    queryset = Financas.objects.all()
    serializer_class = FinancasSerializer

    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)


class FinancasFindByIdView(generics.RetrieveAPIView):
    queryset = Financas.objects.all()
    serializer_class = FinancasSerializer
    lookup_field = 'pk'

    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)


class FinancasFindyByNotionIdView(generics.RetrieveAPIView):
    queryset = Financas.objects.all()
    serializer_class = FinancasSerializer
    lookup_field = 'notion_page_id'

    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)


class FinancasDeleteView(generics.DestroyAPIView):
    queryset = Financas.objects.all()
    serializer_class = FinancasSerializer

    def delete(self, request, *args, **kwargs):
        try:
            # Obter a instância do objeto a ser excluído
            instance = self.get_object()
            notion_page_id = instance.notion_page_id

            # Excluir a página no Notion
            url = f"https://api.notion.com/v1/pages/{notion_page_id}"
            response = requests.patch(
                url, json={"archived": True}, headers=headers)
            if response.status_code != 200:
                logger.error(
                    "Erro ao excluir a página no Notion: %s", response.text)
                raise Exception("Erro ao excluir a página no Notion")
            logger.info(
                "Página com Notion Page ID %s arquivada no Notion.", notion_page_id)

            # Excluir da planilha de Excel
            delete_from_sheet(notion_page_id)

            # Excluir do banco de dados
            instance.delete()
            logger.info(
                "Objeto com Notion Page ID %s excluído do banco de dados.", notion_page_id)

            # Obter detalhes do objeto excluído para retornar na resposta
            serialized_notion = FinancasSerializer(instance).data

            # Registrar o objeto excluído no console
            logger.info("Objeto excluído: %s", json.dumps(
                serialized_notion, indent=4, ensure_ascii=False))

            # Retornar o objeto excluído na resposta da API
            return Response({"message": "Finança excluída com sucesso", "data": serialized_notion}, status=status.HTTP_204_NO_CONTENT)
        except Exception as erro:
            logger.error("Erro ao excluir finança: %s", erro)
            return Response({"message": "Erro ao excluir finança"}, status=status.HTTP_400_BAD_REQUEST)
